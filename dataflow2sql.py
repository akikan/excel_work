#coding:utf-8
import openpyxl as px
import Rect

wb = px.load_workbook('test.xlsx')
ws = wb.get_sheet_by_name('データフロー')


def excelShaping(ws, startY=1, startX=1, endY, endX):
	for y in range(startY, endY):
		for x in range(startX, endX):
			#  ターゲットセルの↓と→のセルを見て隣接部分に罫線があったら両者に罫線をつける
			targetCell=ws.cell(row=y, column=x).border
			rightCell=ws.cell(row=y, column=x+1).border
			bottomCell=ws.cell(row=y+1, column=x).border

			# 右側の罫線確認
			if targetCell.right.style is not None:
				rightCell.left.style = targetCell.right.style
			else if rightCell.left.style is not None:
				targetCell.right.style = rightCell.left.style

			# 下側の罫線確認
			if targetCell.bottom.style is not None:
				bottomCell.top.style = targetCell.bottom.style
			else if bottomCell.left.style is not None:
				targetCell.bottom.style = bottomCell.top.style

	return ws

def getRect(ws, oy,ox, maxHeight, maxWidth):
	width=1
	height=1

	# get rect Width
	for x in range(1, maxWidth):
		if ws.cell(row=oy, column=ox+x).border.right is not None:
			width=x
			break

	# get rect height
	for y in range(1, maxHeight):
		if ws.cell(row=oy+y, column=ox).border.left is None:
			Height=y-1
			break

	# 右下チェック(ここダメなら大抵ダメでしょう。きっと)
	testCell =  ws.cell(row=oy+Height, column=ox+width).border
	if testCell.bottom.style is not None and testCell.right.style is not None:
		return None

	for y in range(oy, height):
		if ws.cell(row=oy+y, column=ox+width-1).border.right is None:
			return None
	return Rect.Rect(oy,ox,height,width)



def searchRect(ws, startY=1, startX=1, endY, endX):
	ret=[]
	for y in range(startY, endY):
		for x in range(startX, endX):
			cell=ws.cell(row=y, column=x).border

			if cell.top.style is not None and cell.left.style is not None:
				ret.add(getRect(ws, y,x, endY, endX))

	return ret

def deleteDuplicateRect(rects):
	ret=[]
	for j in range(length):
		target=rects[j]
		count=0
		for i in range(j+1,length):
			if rects[i].x >= target.x and 
			   rects[i].x+rects[i].width <= target.x+target.width and
			   rects[i].y >= target.y and
			   rects[i].y + rects[i].height <= target.y+target.height:
			   #targetの中に包含されている四角形
			   count+=1
		if count != 0:
			ret.append(target)
	return ret

temp = searchRect(ws, endX=100, endY=100)
rects=deleteDuplicateRect(temp)